import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import openpyxl
from datetime import datetime
import os

class BoothSalesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("매대판매 POS 시스템")
        self.root.geometry("900x700")
        
        # 메모리 데이터 저장소
        self.master_data = {}  
        self.serial_data = {}  
        self.cart = []         
        self.op_file_path = "" 
        self.inspection_dict = {} 
        self.sold_serial_data = {}
        self.discount_data = {}

        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    # 숫자 변환 안정성 확보용 Helper 함수
    def safe_int(self, value):
        if value is None:
            return 0
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        value = str(value).replace(",", "").strip()
        if value == "":
            return 0
        try:
            return int(float(value))
        except ValueError:
            return 0

    def setup_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)

        self.tab_init = ttk.Frame(notebook)
        self.tab_sales = ttk.Frame(notebook)
        self.tab_export = ttk.Frame(notebook)
        self.tab_inspection = ttk.Frame(notebook)
        self.tab_return_exchange = ttk.Frame(notebook) 

        notebook.add(self.tab_init, text="1. 초기설정 (엑셀 로드)")
        notebook.add(self.tab_sales, text="2. 판매 관리")
        notebook.add(self.tab_export, text="3. 리포트 출력")
        notebook.add(self.tab_inspection, text="4. 재고 검수")
        notebook.add(self.tab_return_exchange, text="5. 반품/교환") 

        self.build_init_tab()
        self.build_sales_tab()
        self.build_export_tab()
        self.build_inspection_tab()
        self.build_return_exchange_tab() 

    # ================= 1. 초기 설정 탭 =================
    def build_init_tab(self):
        frame = ttk.LabelFrame(self.tab_init, text="운영 파일 및 기준 정보 설정", padding=20)
        frame.pack(fill='x', padx=20, pady=20)

        ttk.Button(frame, text="1) 기준 엑셀 불러오기 (품목별재고/할인)", command=self.load_master_excel, width=40).grid(row=0, column=0, pady=10, sticky='w')
        self.lbl_master_status = ttk.Label(frame, text="대기 중...", foreground="gray")
        self.lbl_master_status.grid(row=0, column=1, padx=10)

        ttk.Button(frame, text="2) 운영파일 저장위치 지정/생성", command=self.set_op_file, width=40).grid(row=1, column=0, pady=10, sticky='w')
        self.lbl_op_status = ttk.Label(frame, text="대기 중...", foreground="gray")
        self.lbl_op_status.grid(row=1, column=1, padx=10)

        ttk.Button(frame, text="3) 시리얼 엑셀 불러오기 (선택사항)", command=self.load_serial_excel, width=40).grid(row=2, column=0, pady=10, sticky='w')
        self.lbl_serial_status = ttk.Label(frame, text="대기 중...", foreground="gray")
        self.lbl_serial_status.grid(row=2, column=1, padx=10)

    def load_master_excel(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not filepath: return

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "품목별재고" not in wb.sheetnames:
                messagebox.showerror("오류", "'품목별재고' 시트가 존재하지 않습니다.")
                return

            ws = wb["품목별재고"]
            headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}
            
            required_cols = ['바코드', '상품명', '구분', '단가', '재고']
            if not all(col in headers for col in required_cols):
                messagebox.showerror("오류", f"필수 헤더가 누락되었습니다. 필요: {required_cols}")
                return

            self.master_data.clear()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[headers['바코드']]: continue
                
                barcode = str(row[headers['바코드']])
                item_code = ""
                if '상품코드' in headers and row[headers['상품코드']]:
                    item_code = str(row[headers['상품코드']])

                memo = ""
                if '메모' in headers and row[headers['메모']]:
                    memo = str(row[headers['메모']]).strip()

                self.master_data[barcode] = {
                    'item_code': item_code,
                    'name': str(row[headers['상품명']]),
                    'category': str(row[headers['구분']]).strip().upper() if row[headers['구분']] else "ETC",
                    'price': self.safe_int(row[headers['단가']]),
                    'stock': self.safe_int(row[headers['재고']]),
                    'memo': memo
                }

            self.discount_data.clear()
            if "할인" in wb.sheetnames:
                ws_dc = wb["할인"]
                headers_dc = {cell.value: idx for idx, cell in enumerate(ws_dc[1]) if cell.value}
                
                req_dc_cols = ['바코드번호', 'N개이상일시', '할인적용할단가']
                req_dc_cols_space = ['바코드번호', 'N개이상일 시', '할인적용할 단가']
                
                target_cols = req_dc_cols if all(col in headers_dc for col in req_dc_cols) else req_dc_cols_space
                
                if all(col in headers_dc for col in target_cols):
                    for row in ws_dc.iter_rows(min_row=2, values_only=True):
                        bcode = str(row[headers_dc[target_cols[0]]] or "").strip()
                        if not bcode: continue
                        
                        try:
                            min_qty = self.safe_int(row[headers_dc[target_cols[1]]])
                            dc_price = self.safe_int(row[headers_dc[target_cols[2]]])
                            if min_qty > 0:
                                self.discount_data[bcode] = {
                                    'min_qty': min_qty,
                                    'dc_price': dc_price
                                }
                        except ValueError:
                            pass

            dc_msg = f" (할인 정책 {len(self.discount_data)}건 적용됨)" if self.discount_data else ""
            self.lbl_master_status.config(text=f"로드 완료 ({len(self.master_data)}개 품목){dc_msg}", foreground="blue")
            messagebox.showinfo("성공", f"상품 마스터를 성공적으로 불러왔습니다.{dc_msg}")
            
        except Exception as e:
            messagebox.showerror("오류", f"기준 엑셀 로드 실패:\n{e}")

    def set_op_file(self):
        if not self.master_data:
            messagebox.showerror("순서 오류", "1번 '기준 엑셀 불러오기'를 먼저 실행해야 합니다.")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")], 
            initialfile="운영파일.xlsx",
            title="운영파일을 선택하거나 새 이름을 입력하세요"
        )
        if not filepath: return
        
        self.op_file_path = filepath
        
        if os.path.exists(self.op_file_path):
            try:
                wb = openpyxl.load_workbook(self.op_file_path)
                modified_for_compat = False
                
                if "판매상세" in wb.sheetnames:
                    ws_sales = wb["판매상세"]
                    if ws_sales.max_column < 9 or ws_sales.cell(row=1, column=9).value != "비고":
                        ws_sales.cell(row=1, column=9, value="비고")
                        modified_for_compat = True

                if "상품마스터" in wb.sheetnames:
                    ws_master = wb["상품마스터"]
                    master_headers = {cell.value: idx for idx, cell in enumerate(ws_master[1], start=1) if cell.value}
                    if "메모" not in master_headers:
                        ws_master.cell(row=1, column=ws_master.max_column + 1, value="메모")
                        modified_for_compat = True
                
                if "반품교환상세" not in wb.sheetnames:
                    ws_ret = wb.create_sheet("반품교환상세")
                    ws_ret.append(["처리일자", "시간", "처리구분", "원판매일자", "반품바코드", "반품상품명", "반품수량", "반품단가", "반품금액", "반품시리얼코드", "교환출고바코드", "교환출고상품명", "교환출고수량", "교환출고단가", "교환출고금액", "교환출고시리얼코드", "차액", "처리메모"])
                    modified_for_compat = True

                if "단가변경이력" not in wb.sheetnames:
                    ws_price_hist = wb.create_sheet("단가변경이력")
                    ws_price_hist.append(["변경일자", "시간", "변경기준", "입력상품코드", "바코드", "상품코드", "상품명", "이전단가", "변경단가", "변경메모"])
                    modified_for_compat = True
                    
                if modified_for_compat:
                    wb.save(self.op_file_path)

                # 💡 버그 수정: 운영파일 "상품마스터" 데이터를 무조건 우선시하여 master_data 덮어쓰기
                if "상품마스터" in wb.sheetnames:
                    ws_master = wb["상품마스터"]
                    master_headers = {cell.value: idx for idx, cell in enumerate(ws_master[1]) if cell.value}
                    for row in ws_master.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            bcode = str(row[0]).strip()
                            name = str(row[master_headers.get("상품명", 1)] or "")
                            category = str(row[master_headers.get("구분", 2)] or "ETC").strip().upper()
                            price = self.safe_int(row[master_headers.get("단가", 3)])
                            item_code = str(row[master_headers.get("상품코드", 5)] or "").strip() if "상품코드" in master_headers else ""
                            old_memo = self.master_data[bcode].get("memo", "") if bcode in self.master_data else ""
                            memo = str(row[master_headers["메모"]] or old_memo).strip() if "메모" in master_headers else old_memo

                            # 기존 재고 보존 (뒤에 재고현황 시트에서 덮어씀)
                            old_stock = self.master_data[bcode]["stock"] if bcode in self.master_data else 0

                            self.master_data[bcode] = {
                                "item_code": item_code,
                                "name": name,
                                "category": category,
                                "price": price,
                                "stock": old_stock,
                                "memo": memo
                            }

                if "재고현황" in wb.sheetnames:
                    ws_stock = wb["재고현황"]
                    stock_count = 0
                    for row in ws_stock.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            bcode = str(row[0])
                            if bcode in self.master_data:
                                self.master_data[bcode]['stock'] = self.safe_int(row[2])
                                stock_count += 1
                
                if "시리얼목록" in wb.sheetnames:
                    ws_serial = wb["시리얼목록"]
                    self.serial_data.clear()
                    self.sold_serial_data.clear()
                    
                    for row in ws_serial.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[2]:
                            bcode, scode, status = str(row[0]), str(row[2]), row[3]
                            if status != "O":
                                if bcode not in self.serial_data:
                                    self.serial_data[bcode] = []
                                self.serial_data[bcode].append(scode)
                            else:
                                if bcode not in self.sold_serial_data:
                                    self.sold_serial_data[bcode] = set()
                                self.sold_serial_data[bcode].add(scode)

                self.lbl_op_status.config(text=f"연결됨: {os.path.basename(self.op_file_path)}", foreground="blue")
                messagebox.showinfo("성공", "기존 운영파일을 불러와 최신 단가, 재고, 시리얼 정보를 성공적으로 동기화했습니다.")
                return 

            except Exception as e:
                messagebox.showerror("오류", f"기존 파일 로드 중 오류 발생: {e}")
                return

        # 3. 파일이 없는 경우 (신규 생성)
        try:
            wb = openpyxl.Workbook()
            sheets = ["상품마스터", "판매상세", "일별집계", "시리얼목록", "재고현황", "반품교환상세", "단가변경이력"]
            for i, name in enumerate(sheets):
                if i == 0: ws = wb.active; ws.title = name
                else: ws = wb.create_sheet(name)

            wb["상품마스터"].append(["바코드", "상품명", "구분", "단가", "초기재고", "상품코드", "메모"])
            wb["판매상세"].append(["판매일자", "시간", "바코드", "상품명", "수량", "단가", "총액", "시리얼코드", "비고"])
            wb["일별집계"].append(["판매일자", "바코드", "상품명", "판매수량합계", "판매금액합계"])
            wb["시리얼목록"].append(["바코드", "상품명", "시리얼번호", "판매여부"])
            wb["재고현황"].append(["바코드", "상품명", "현재재고"])
            wb["반품교환상세"].append(["처리일자", "시간", "처리구분", "원판매일자", "반품바코드", "반품상품명", "반품수량", "반품단가", "반품금액", "반품시리얼코드", "교환출고바코드", "교환출고상품명", "교환출고수량", "교환출고단가", "교환출고금액", "교환출고시리얼코드", "차액", "처리메모"])
            wb["단가변경이력"].append(["변경일자", "시간", "변경기준", "입력상품코드", "바코드", "상품코드", "상품명", "이전단가", "변경단가", "변경메모"])
            
            if self.master_data:
                for bcode, info in self.master_data.items():
                    wb["상품마스터"].append([bcode, info['name'], info['category'], info['price'], info['stock'], info.get('item_code', ''), info.get('memo', '')])
                    wb["재고현황"].append([bcode, info['name'], info['stock']])

            if self.serial_data:
                for bcode, serials in self.serial_data.items():
                    name = self.master_data[bcode]['name'] if bcode in self.master_data else "알수없음"
                    for s in serials:
                        wb["시리얼목록"].append([bcode, name, s, ""])

            wb.save(self.op_file_path)
            self.lbl_op_status.config(text=f"신규 생성: {os.path.basename(self.op_file_path)}", foreground="darkgreen")
            messagebox.showinfo("성공", "새로운 운영파일이 생성되고 초기 데이터가 기록되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"운영파일 생성 중 오류 발생: {e}")

    def load_serial_excel(self):
        if not self.master_data:
            messagebox.showerror("순서 오류", "1번 '기준 엑셀 불러오기'를 먼저 실행해야 합니다.")
            return

        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not filepath: return

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}

            if 'barcode' not in headers or 'serial_code' not in headers:
                messagebox.showerror("오류", "첫 행에 'barcode', 'serial_code' 헤더가 필요합니다.")
                return

            existing_op_serials = set()
            op_wb = None
            ws_serial_op = None
            
            if self.op_file_path and os.path.exists(self.op_file_path):
                op_wb = openpyxl.load_workbook(self.op_file_path)
                if "시리얼목록" in op_wb.sheetnames:
                    ws_serial_op = op_wb["시리얼목록"]
                    for row in ws_serial_op.iter_rows(min_row=2, values_only=True):
                        if row[2]:
                            existing_op_serials.add(str(row[2]))

            new_serials_to_add = []
            count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                bcode = str(row[headers['barcode']])
                scode = str(row[headers['serial_code']])
                
                if bcode and scode:
                    if scode in existing_op_serials:
                        continue
                        
                    if bcode not in self.serial_data:
                        self.serial_data[bcode] = []
                    
                    if scode not in self.serial_data[bcode]:
                        self.serial_data[bcode].append(scode)
                        new_serials_to_add.append((bcode, scode))
                        existing_op_serials.add(scode)
                        count += 1

            if op_wb and ws_serial_op and new_serials_to_add:
                for bcode, scode in new_serials_to_add:
                    name = self.master_data[bcode]['name'] if bcode in self.master_data else "알수없음"
                    ws_serial_op.append([bcode, name, scode, ""])
                
                op_wb.save(self.op_file_path)

            self.lbl_serial_status.config(text=f"로드 완료 (신규 {count}개 추가)", foreground="blue")
            
            if count > 0:
                messagebox.showinfo("성공", f"새로운 시리얼 {count}건을 성공적으로 추가했습니다.")
            else:
                messagebox.showinfo("안내", "추가할 신규 시리얼이 없습니다.\n(모두 이미 등록되었거나 판매된 시리얼입니다.)")
            
        except PermissionError:
            messagebox.showerror("오류", "운영파일이 엑셀에서 열려 있습니다. 파일을 닫고 다시 시도해 주세요.")
        except Exception as e:
            messagebox.showerror("오류", f"시리얼 엑셀 로드 실패:\n{e}")

    # ================= 2. 판매 관리 탭 =================
    def build_sales_tab(self):
        scan_frame = ttk.LabelFrame(self.tab_sales, text="바코드 및 시리얼 스캔", padding=10)
        scan_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(scan_frame, text="바코드:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_barcode = ttk.Entry(scan_frame, width=20)
        self.entry_barcode.grid(row=0, column=1, padx=5, pady=5)
        self.entry_barcode.bind('<Return>', self.handle_barcode_enter)

        ttk.Label(scan_frame, text="시리얼 (필수상품용):").grid(row=0, column=2, padx=5, pady=5)
        self.entry_serial = ttk.Entry(scan_frame, width=20)
        self.entry_serial.grid(row=0, column=3, padx=5, pady=5)
        self.entry_serial.bind('<Return>', lambda e: self.add_to_cart())

        ttk.Button(scan_frame, text="장바구니 추가", command=self.add_to_cart).grid(row=0, column=4, padx=10)

        cart_frame = ttk.Frame(self.tab_sales)
        cart_frame.pack(fill='both', expand=True, padx=10, pady=5)

        cols = ("바코드", "상품명", "수량", "현재재고", "단가", "총액", "시리얼코드", "메모", "삭제")
        self.tree_cart = ttk.Treeview(cart_frame, columns=cols, show='headings')
        for col in cols:
            self.tree_cart.heading(col, text=col)
            if col == "삭제": width = 50
            elif col in ["수량", "현재재고"]: width = 70
            elif col == "메모": width = 140
            else: width = 100
            self.tree_cart.column(col, width=width, anchor='center')
        self.tree_cart.pack(fill='both', expand=True)

        self.tree_cart.bind('<ButtonRelease-1>', self.on_cart_click)        
        self.tree_cart.bind('<Double-1>', self.on_cart_double_click)

        pay_frame = ttk.Frame(self.tab_sales)
        pay_frame.pack(fill='x', padx=10, pady=10)

        self.lbl_total = ttk.Label(pay_frame, text="총 결제금액: 0 원", font=("Arial", 14, "bold"))
        self.lbl_total.pack(side='left', padx=10)

        self.btn_complete = ttk.Button(pay_frame, text="판매완료", command=self.complete_sale, style="Accent.TButton", width=15)
        self.btn_complete.pack(side='right', padx=10)

        ttk.Button(pay_frame, text="장바구니 비우기", command=self.clear_cart).pack(side='right')

    def add_to_cart(self, event=None):
        if not self.op_file_path or not self.master_data:
            messagebox.showwarning("경고", "초기설정 탭에서 기준 엑셀과 운영파일을 먼저 설정하세요.")
            return

        barcode = self.entry_barcode.get().strip()
        serial = self.entry_serial.get().strip()

        if not barcode: return

        if barcode not in self.master_data:
            messagebox.showerror("오류", f"등록되지 않은 바코드입니다: {barcode}")
            self.entry_barcode.delete(0, tk.END)
            self.entry_serial.delete(0, tk.END)
            self.entry_barcode.focus()
            return

        item = self.master_data[barcode]
        is_serial_req = item['category'] in ['RUBBER', 'RACKET']

        if is_serial_req:
            if not serial:
                messagebox.showwarning("입력 필요", f"[{item['name']}] 상품은 시리얼 번호 스캔이 필요합니다.")
                self.entry_serial.focus()
                return
            
            if hasattr(self, 'sold_serial_data') and barcode in self.sold_serial_data:
                if serial in self.sold_serial_data[barcode]:
                    messagebox.showerror("오류", "이미 판매가 완료되어 출고된 시리얼 번호입니다!")
                    self.entry_serial.delete(0, tk.END)
                    self.entry_serial.focus()
                    return

            for c_item in self.cart:
                if c_item['serial'] == serial:
                    messagebox.showerror("오류", "이미 장바구니에 담긴 시리얼 번호입니다.")
                    self.entry_serial.delete(0, tk.END)
                    return

        current_cart_qty = sum(c['qty'] for c in self.cart if c['barcode'] == barcode)
        if item['stock'] < current_cart_qty + 1:
            messagebox.showerror("재고 부족", f"현재 남은 전산 재고는 {item['stock']}개 입니다.")
            self.entry_barcode.delete(0, tk.END)
            self.entry_serial.delete(0, tk.END)
            self.entry_barcode.focus()
            return

        if is_serial_req:
            self.cart.append({
                'barcode': barcode, 
                'name': item['name'], 
                'qty': 1, 
                'price': item['price'], 
                'serial': serial,
                'memo': item.get('memo', ''),
                'manual_price': False,
                'original_base_price': item['price'] 
            })
        else:
            found = False
            for c_item in self.cart:
                if c_item['barcode'] == barcode:
                    c_item['qty'] += 1
                    found = True
                    break
            if not found:
                self.cart.append({
                    'barcode': barcode, 
                    'name': item['name'], 
                    'qty': 1, 
                    'price': item['price'], 
                    'serial': '',
                    'memo': item.get('memo', ''),
                    'manual_price': False,
                    'original_base_price': item['price'] 
                })

        self.update_cart_ui()
        self.entry_barcode.delete(0, tk.END)
        self.entry_serial.delete(0, tk.END)
        self.entry_barcode.focus()

    def update_cart_ui(self):
        cart_qty_by_barcode = {}
        for item in self.cart:
            bcode = item['barcode']
            cart_qty_by_barcode[bcode] = cart_qty_by_barcode.get(bcode, 0) + item['qty']
            
        for item in self.cart:
            if item.get('manual_price'): 
                continue 
                
            bcode = item['barcode']
            base_price = item.get('original_base_price', item['price'])
            
            if hasattr(self, 'discount_data') and bcode in self.discount_data:
                dc_info = self.discount_data[bcode]
                if cart_qty_by_barcode[bcode] >= dc_info['min_qty']:
                    item['price'] = dc_info['dc_price']
                else:
                    item['price'] = base_price
            else:
                item['price'] = base_price

        for row in self.tree_cart.get_children():
            self.tree_cart.delete(row)
        
        total_amt = 0
        for item in self.cart:
            amt = item['qty'] * item['price']
            total_amt += amt
            
            current_stock = 0
            if item['barcode'] in self.master_data:
                current_stock = self.master_data[item['barcode']]['stock']

            self.tree_cart.insert("", tk.END, values=(
                item['barcode'], item['name'], item['qty'], current_stock, 
                item['price'], amt, item['serial'], item.get('memo', ''), "❌"
            ))
            
        self.lbl_total.config(text=f"총 결제금액: {total_amt:,} 원")

    def clear_cart(self):
        self.cart.clear()
        self.update_cart_ui()

    def complete_sale(self):
        if not self.cart:
            messagebox.showwarning("알림", "장바구니가 비어 있습니다.")
            return
        
        if hasattr(self, 'btn_complete'):
            self.btn_complete.config(state='disabled')

        try:
            wb = openpyxl.load_workbook(self.op_file_path)
            ws_sales = wb["판매상세"]
            ws_stock = wb["재고현황"]
            ws_serial = wb["시리목록"] if "시리얼목록" not in wb.sheetnames else wb["시리얼목록"]
            ws_daily = wb["일별집계"]

            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")

            for item in self.cart:
                barcode = str(item['barcode'])
                qty = int(item['qty'])
                price = int(item['price'])
                amt = qty * price
                serial = str(item['serial']) if item['serial'] else ""

                ws_sales.append([date_str, time_str, barcode, item['name'], qty, price, amt, serial, item.get('memo', '')])

                found_daily = False
                for row in ws_daily.iter_rows(min_row=2):
                    sale_date_cell = str(row[0].value)[:10] if row[0].value else ""
                    if sale_date_cell == date_str and str(row[1].value) == barcode:
                        row[3].value = int(row[3].value or 0) + qty
                        row[4].value = int(row[4].value or 0) + amt
                        found_daily = True
                        break
                        
                if not found_daily:
                    ws_daily.append([date_str, barcode, item['name'], qty, amt])

                if serial:
                    serial_found_in_op = False
                    for row in ws_serial.iter_rows(min_row=2):
                        if str(row[0].value) == barcode and str(row[2].value) == serial:
                            row[3].value = "O"
                            serial_found_in_op = True
                            break
                    
                    if not serial_found_in_op:
                        ws_serial.append([barcode, item['name'], serial, "O"])

                    if barcode in self.serial_data and serial in self.serial_data[barcode]:
                        self.serial_data[barcode].remove(serial)
                        
                    if not hasattr(self, 'sold_serial_data'):
                        self.sold_serial_data = {}
                    if barcode not in self.sold_serial_data:
                        self.sold_serial_data[barcode] = set()
                    self.sold_serial_data[barcode].add(serial)

                self.master_data[barcode]['stock'] -= qty

            ws_stock.delete_rows(2, ws_stock.max_row)
            for b_code, b_data in self.master_data.items():
                ws_stock.append([b_code, b_data['name'], b_data['stock']])

            wb.save(self.op_file_path)
            messagebox.showinfo("완료", "판매 기록이 완료되었습니다.\n(미등록 시리얼이 있었다면 자동으로 등록되었습니다.)")
            self.clear_cart()

        except PermissionError:
            for item in self.cart:
                bcode = str(item['barcode'])
                qty = int(item['qty'])
                serial = str(item['serial']) if item['serial'] else ""
                
                if bcode in self.master_data:
                    self.master_data[bcode]['stock'] += qty
                if serial:
                    if bcode in self.serial_data:
                        self.serial_data[bcode].append(serial)
                    if hasattr(self, 'sold_serial_data') and bcode in self.sold_serial_data:
                        if serial in self.sold_serial_data[bcode]:
                            self.sold_serial_data[bcode].remove(serial)
                            
            messagebox.showerror("오류", "운영파일이 엑셀 등에서 열려 있습니다. 파일을 닫고 다시 [판매완료]를 눌러주세요.")
            
        except Exception as e:
            for item in self.cart:
                bcode = str(item['barcode'])
                qty = int(item['qty'])
                serial = str(item['serial']) if item['serial'] else ""
                
                if bcode in self.master_data:
                    self.master_data[bcode]['stock'] += qty
                if serial:
                    if bcode in self.serial_data:
                        self.serial_data[bcode].append(serial)
                    if hasattr(self, 'sold_serial_data') and bcode in self.sold_serial_data:
                        if serial in self.sold_serial_data[bcode]:
                            self.sold_serial_data[bcode].remove(serial)
                            
            messagebox.showerror("오류", f"저장 중 오류 발생 (데이터가 안전하게 복구되었습니다):\n{e}")
            
        finally:
            if hasattr(self, 'btn_complete'):
                self.btn_complete.config(state='normal')

    def on_cart_click(self, event):
        region = self.tree_cart.identify_region(event.x, event.y)
        if region != "cell": return
            
        column = self.tree_cart.identify_column(event.x)
        row_id = self.tree_cart.identify_row(event.y)
        
        if column == '#9' and row_id:
            values = self.tree_cart.item(row_id, 'values')
            clicked_barcode = str(values[0])
            clicked_serial = str(values[6]) if values[6] and values[6] != 'None' else ""
            
            for i, item in enumerate(self.cart):
                item_serial = str(item['serial']) if item['serial'] else ""
                if str(item['barcode']) == clicked_barcode and item_serial == clicked_serial:
                    del self.cart[i]
                    break
                    
            self.update_cart_ui()

    def on_cart_double_click(self, event):
        region = self.tree_cart.identify_region(event.x, event.y)
        if region != "cell": return
            
        column = self.tree_cart.identify_column(event.x)
        row_id = self.tree_cart.identify_row(event.y)
        
        if column == '#5' and row_id:
            values = self.tree_cart.item(row_id, 'values')
            clicked_barcode = str(values[0])
            item_name = str(values[1])
            clicked_serial = str(values[6]) if values[6] and values[6] != 'None' else ""
            
            current_price = int(values[4])
            
            new_price = simpledialog.askinteger(
                "단가 수정", 
                f"[{item_name}]의 새로운 단가를 입력하세요:", 
                initialvalue=current_price, 
                parent=self.root,
                minvalue=0
            )
            
            if new_price is not None:
                for item in self.cart:
                    item_serial = str(item['serial']) if item['serial'] else ""
                    if str(item['barcode']) == clicked_barcode and item_serial == clicked_serial:
                        item['price'] = new_price
                        item['manual_price'] = True 
                        break
                        
                self.update_cart_ui()

    def handle_barcode_enter(self, event):
        barcode = self.entry_barcode.get().strip()
        if not barcode: return

        if barcode not in self.master_data:
            messagebox.showerror("오류", f"등록되지 않은 바코드입니다: {barcode}")
            self.entry_barcode.delete(0, tk.END)
            return

        item = self.master_data[barcode]
        is_serial_req = item['category'] in ['RUBBER', 'RACKET']

        if is_serial_req:
            self.entry_serial.focus()
        else:
            self.add_to_cart()

    def on_closing(self):
        if messagebox.askyesno("확인", "판매일보를 생성하셨나요?\n(생성하지 않고 종료하면 오늘 판매분 집계가 어려울 수 있습니다.)"):
            if messagebox.askyesno("종료", "정말로 프로그램을 종료하시겠습니까?"):
                self.root.destroy()
        else:
            messagebox.showinfo("안내", "리포트 출력 탭에서 판매일보를 생성해 주세요.")

    # ================= 3. 리포트 출력 탭 =================
    def build_export_tab(self):
        frame = ttk.LabelFrame(self.tab_export, text="데이터 내보내기", padding=20)
        frame.pack(fill='x', padx=20, pady=20)

        self.lbl_daily_total = ttk.Label(frame, text="오늘의 총 판매/환불 합계액: 0 원", font=("Arial", 12, "bold"), foreground="blue")
        self.lbl_daily_total.grid(row=0, column=0, columnspan=2, pady=10, sticky='w')

        ttk.Button(frame, text="판매일보 생성 및 내보내기", command=self.export_daily_report).grid(row=1, column=0, pady=10, padx=10, sticky='w')
        ttk.Label(frame, text="* 판매상세(반품/비고 포함) 및 품목별 집계가 포함된 새로운 엑셀을 생성합니다.", foreground="gray").grid(row=1, column=1, sticky='w')

    def export_daily_report(self):
        if not self.op_file_path or not os.path.exists(self.op_file_path):
            messagebox.showerror("오류", "운영 파일이 지정되지 않았거나 없습니다.")
            return

        today_str = datetime.now().strftime("%Y-%m-%d")

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")], 
            initialfile=f"판매일보_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not save_path: return

        try:
            wb_op = openpyxl.load_workbook(self.op_file_path, data_only=True)
            ws_sales = wb_op["판매상세"]

            wb_report = openpyxl.Workbook()
            ws_report_sales = wb_report.active
            ws_report_sales.title = "판매상세"
            ws_summary = wb_report.create_sheet("품목별집계")

            summary_data = {} 
            grand_total = 0  
            today_sales_count = 0 
            
            headers = {cell.value: idx for idx, cell in enumerate(ws_sales[1]) if cell.value}
            idx_date = headers.get("판매일자", 0)
            idx_bcode = headers.get("바코드", 2)
            idx_name = headers.get("상품명", 3)
            idx_qty = headers.get("수량", 4)
            idx_amt = headers.get("총액", 6)

            for i, row in enumerate(ws_sales.iter_rows(values_only=True)):
                row_list = list(row)
                
                if i == 0: 
                    row_list.insert(3, "상품코드")
                    if len(row_list) < 10: 
                        while len(row_list) < 10:
                            row_list.append("")
                        row_list[9] = "비고"
                    ws_report_sales.append(row_list)
                    continue 
                
                raw_date = row[idx_date]
                if isinstance(raw_date, datetime):
                    sale_date = raw_date.strftime("%Y-%m-%d")
                else:
                    sale_date = str(raw_date)[:10] if raw_date else ""
                
                if sale_date != today_str:
                    continue

                today_sales_count += 1
                
                bcode = str(row[idx_bcode]) if row[idx_bcode] is not None else ""
                name = row[idx_name] if row[idx_name] is not None else ""
                
                qty = self.safe_int(row[idx_qty])
                amt = self.safe_int(row[idx_amt])
                
                # 판매금액은 현재 상품마스터 단가로 재계산하지 않고, 판매상세에 기록된 총액을 그대로 합산한다. 단가 변경 후에도 과거 판매금액을 보존하기 위함.
                grand_total += amt

                item_code = ""
                if bcode in self.master_data:
                    item_code = self.master_data[bcode].get('item_code', '')

                row_list.insert(3, item_code)
                ws_report_sales.append(row_list)

                if bcode not in summary_data:
                    summary_data[bcode] = {'item_code': item_code, 'name': name, 'qty': 0, 'amt': 0}
                
                summary_data[bcode]['qty'] += qty
                summary_data[bcode]['amt'] += amt

            if today_sales_count == 0:
                messagebox.showwarning("안내", f"오늘({today_str}) 판매/반품 내역이 없습니다.")

            ws_summary.append(["바코드", "상품코드", "상품명", "순판매수량", "순판매금액"])
            for bcode, data in summary_data.items():
                ws_summary.append([bcode, data['item_code'], data['name'], data['qty'], data['amt']])
            
            ws_summary.append([])
            ws_summary.append(["", "", "오늘의 순매출 총액", "", grand_total])

            self.lbl_daily_total.config(text=f"오늘의 총 판매/환불 합계액: {grand_total:,} 원")

            wb_report.save(save_path)
            messagebox.showinfo("성공", f"오늘의 판매일보가 저장되었습니다.\n순매출 총액: {grand_total:,}원")

        except Exception as e:
            messagebox.showerror("오류", f"리포트 생성 실패:\n{e}")

    # ================= 4. 재고 검수 탭 =================
    def build_inspection_tab(self):
        top_frame = ttk.Frame(self.tab_inspection)
        top_frame.pack(fill='x', padx=10, pady=10)

        scan_frame = ttk.LabelFrame(top_frame, text="바코드 스캔", padding=10)
        scan_frame.pack(side='left', fill='x', expand=True, padx=(0, 10))

        ttk.Label(scan_frame, text="바코드:").pack(side='left', padx=5)
        self.entry_inspect_barcode = ttk.Entry(scan_frame, width=30)
        self.entry_inspect_barcode.pack(side='left', padx=5)
        self.entry_inspect_barcode.bind('<Return>', self.handle_inspect_barcode)

        btn_frame = ttk.Frame(top_frame)
        btn_frame.pack(side='right')

        ttk.Button(btn_frame, text="신규 상품 등록", command=self.open_add_product_window).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="상품코드 단가 변경", command=self.open_change_price_by_itemcode_window).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="전체 품목 불러오기", command=self.load_all_for_inspection).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="초기화", command=self.clear_inspection).pack(side='left', padx=5)

        list_frame = ttk.Frame(self.tab_inspection)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        cols = ("바코드", "상품코드", "상품명", "전산재고", "실물재고", "차이(실물-전산)", "상태")
        self.tree_inspect = ttk.Treeview(list_frame, columns=cols, show='headings')
        
        for col in cols:
            self.tree_inspect.heading(col, text=col)
            width = 80 if "재고" in col or col == "차이(실물-전산)" else 120
            if col == "상품명": width = 200
            self.tree_inspect.column(col, width=width, anchor='center')
        
        self.tree_inspect.pack(fill='both', expand=True)
        self.tree_inspect.bind('<Double-1>', self.on_inspect_double_click)

        self.tree_inspect.tag_configure('match', foreground='blue')
        self.tree_inspect.tag_configure('mismatch', foreground='red')

        bottom_frame = ttk.Frame(self.tab_inspection)
        bottom_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(bottom_frame, text="* 팁: '전산재고'를 더블클릭하면 추가입고 반영을, '실물재고'를 더블클릭하면 검수수량을 수정할 수 있습니다.", foreground="gray").pack(side='left')
        ttk.Button(bottom_frame, text="검수결과 엑셀 내보내기", command=self.export_inspection_report, style="Accent.TButton").pack(side='right')

    def handle_inspect_barcode(self, event):
        if not self.master_data:
            messagebox.showerror("오류", "먼저 1번 탭에서 기준 엑셀을 불러와주세요.")
            return

        barcode = self.entry_inspect_barcode.get().strip()
        if not barcode: return

        if barcode not in self.master_data:
            messagebox.showerror("오류", f"등록되지 않은 바코드입니다: {barcode}")
            self.entry_inspect_barcode.delete(0, tk.END)
            return

        if barcode not in self.inspection_dict:
            self.inspection_dict[barcode] = 0
        
        self.inspection_dict[barcode] += 1
        
        self.update_inspection_ui()
        self.entry_inspect_barcode.delete(0, tk.END)

    def load_all_for_inspection(self):
        if not self.master_data:
            messagebox.showerror("오류", "기준 데이터가 없습니다.")
            return
            
        if messagebox.askyesno("확인", "등록된 전체 품목을 리스트에 추가하시겠습니까?\n(기존에 스캔한 데이터는 유지됩니다.)"):
            for barcode in self.master_data:
                if barcode not in self.inspection_dict:
                    self.inspection_dict[barcode] = 0
            self.update_inspection_ui()

    def update_inspection_ui(self):
        for row in self.tree_inspect.get_children():
            self.tree_inspect.delete(row)

        for barcode, physical_qty in self.inspection_dict.items():
            info = self.master_data[barcode]
            system_qty = info['stock']
            diff = physical_qty - system_qty
            
            if diff == 0:
                status = "일치"
                tag = 'match'
            elif diff > 0:
                status = "초과"
                tag = 'mismatch'
            else:
                status = "부족"
                tag = 'mismatch'

            self.tree_inspect.insert("", tk.END, values=(
                barcode, info.get('item_code', ''), info['name'], 
                system_qty, physical_qty, diff, status
            ), tags=(tag,))

    def on_inspect_double_click(self, event):
        region = self.tree_inspect.identify_region(event.x, event.y)
        if region != "cell": return
            
        column = self.tree_inspect.identify_column(event.x)
        row_id = self.tree_inspect.identify_row(event.y)
        
        if column == '#5' and row_id:
            values = self.tree_inspect.item(row_id, 'values')
            barcode = str(values[0])
            item_name = str(values[2])
            current_qty = int(values[4])
            
            new_qty = simpledialog.askinteger(
                "실물 수량 수정", 
                f"[{item_name}]의 정확한 실물 수량을 입력하세요:", 
                initialvalue=current_qty, 
                parent=self.root,
                minvalue=0
            )
            
            if new_qty is not None:
                self.inspection_dict[barcode] = new_qty
                self.update_inspection_ui()
                
        elif column == '#4' and row_id:
            values = self.tree_inspect.item(row_id, 'values')
            barcode = str(values[0])
            item_name = str(values[2])
            current_sys_qty = int(values[3])
            
            new_sys_qty = simpledialog.askinteger(
                "전산 재고(입고) 수정", 
                f"[{item_name}]의 현재 총 전산 재고를 입력하세요:\n(기존: {current_sys_qty}개)", 
                initialvalue=current_sys_qty, 
                parent=self.root,
                minvalue=0
            )
            
            if new_sys_qty is not None:
                self.master_data[barcode]['stock'] = new_sys_qty
                
                if self.sync_master_to_op_file():
                    self.update_inspection_ui()
                    messagebox.showinfo("성공", "전산 재고가 수정되어 엑셀 운영파일에 반영되었습니다.")
                else:
                    self.master_data[barcode]['stock'] = current_sys_qty

    def clear_inspection(self):
        if messagebox.askyesno("확인", "현재까지 검수한 실물 재고 데이터를 모두 초기화하시겠습니까?"):
            self.inspection_dict.clear()
            self.update_inspection_ui()

    def export_inspection_report(self):
        if not self.inspection_dict:
            messagebox.showwarning("경고", "내보낼 검수 데이터가 없습니다.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")], 
            initialfile=f"재고실사보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not save_path: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "재고검수결과"

            ws.append(["바코드", "상품코드", "상품명", "전산재고", "실물재고", "차이수량", "상태"])
            
            for barcode, physical_qty in self.inspection_dict.items():
                info = self.master_data[barcode]
                system_qty = info['stock']
                diff = physical_qty - system_qty
                
                status = "일치" if diff == 0 else ("초과" if diff > 0 else "부족")

                ws.append([
                    barcode, info.get('item_code', ''), info['name'], 
                    system_qty, physical_qty, diff, status
                ])

            wb.save(save_path)
            messagebox.showinfo("성공", "재고 실사 보고서가 성공적으로 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"보고서 저장 실패:\n{e}")

    def sync_master_to_op_file(self):
        if not self.op_file_path or not os.path.exists(self.op_file_path):
            messagebox.showwarning("경고", "데이터를 저장할 '운영파일'이 설정되지 않았습니다.\n1번 탭에서 2) 운영파일 지정을 먼저 완료해 주세요.")
            return False 

        try:
            wb = openpyxl.load_workbook(self.op_file_path)
            
            if "재고현황" in wb.sheetnames:
                ws_stock = wb["재고현황"]
                ws_stock.delete_rows(2, ws_stock.max_row)
                for b_code, b_data in self.master_data.items():
                    ws_stock.append([b_code, b_data['name'], b_data['stock']])
                    
            if "상품마스터" in wb.sheetnames:
                ws_master = wb["상품마스터"]
                master_headers = {cell.value: idx for idx, cell in enumerate(ws_master[1], start=1) if cell.value}
                if "메모" not in master_headers:
                    ws_master.cell(row=1, column=ws_master.max_column + 1, value="메모")
                ws_master.delete_rows(2, ws_master.max_row)
                for b_code, b_data in self.master_data.items():
                    ws_master.append([b_code, b_data['name'], b_data['category'], b_data['price'], b_data['stock'], b_data.get('item_code', ''), b_data.get('memo', '')])

            wb.save(self.op_file_path)
            return True
        except Exception as e:
            messagebox.showerror("동기화 오류", f"엑셀 파일에 저장하지 못했습니다. 파일이 열려있는지 확인하세요.\n{e}")
            return False

    def open_add_product_window(self):
        if not self.master_data:
            messagebox.showwarning("경고", "먼저 1번 탭에서 기준 엑셀을 불러오세요.")
            return

        win = tk.Toplevel(self.root)
        win.title("신규 상품 등록")
        win.geometry("400x430")
        win.grab_set() 

        entries = {}
        labels = ['바코드 (필수)', '상품코드', '상품명 (필수)', '구분 (RACKET/RUBBER/ETC)', '단가', '현재 재고', '메모']
        
        for i, text in enumerate(labels):
            ttk.Label(win, text=text).grid(row=i, column=0, padx=15, pady=15, sticky='w')
            e = ttk.Entry(win, width=25)
            e.grid(row=i, column=1, padx=15, pady=15)
            entries[text] = e

        def save_new_product():
            bcode = entries['바코드 (필수)'].get().strip()
            icode = entries['상품코드'].get().strip()
            name = entries['상품명 (필수)'].get().strip()
            category = entries['구분 (RACKET/RUBBER/ETC)'].get().strip() or "ETC"
            memo = entries['메모'].get().strip()
            
            try:
                price = int(entries['단가'].get().strip() or 0)
                stock = int(entries['현재 재고'].get().strip() or 0)
            except ValueError:
                messagebox.showerror("오류", "단가와 재고는 숫자로만 입력해주세요.", parent=win)
                return

            if not bcode or not name:
                messagebox.showerror("오류", "바코드와 상품명은 반드시 입력해야 합니다.", parent=win)
                return

            if bcode in self.master_data:
                messagebox.showerror("오류", "이미 등록된 바코드입니다.", parent=win)
                return

            self.master_data[bcode] = {
                'item_code': icode,
                'name': name,
                'category': category.upper(),
                'price': price,
                'stock': stock,
                'memo': memo
            }
            
            self.inspection_dict[bcode] = 0

            if self.sync_master_to_op_file():
                self.update_inspection_ui()
                messagebox.showinfo("성공", "신규 상품이 등록 및 엑셀에 저장되었습니다.", parent=win)
                win.destroy()
            else:
                if bcode in self.master_data: del self.master_data[bcode]
                if bcode in self.inspection_dict: del self.inspection_dict[bcode]

        ttk.Button(win, text="상품 저장", command=save_new_product, style="Accent.TButton").grid(row=len(labels), column=0, columnspan=2, pady=20)

    def open_change_price_by_itemcode_window(self):
        if not self.master_data:
            messagebox.showwarning("경고", "먼저 1번 탭에서 기준 엑셀을 불러오세요.")
            return

        if not self.op_file_path or not os.path.exists(self.op_file_path):
            messagebox.showwarning("경고", "운영 파일이 지정되지 않았거나 없습니다.")
            return

        win = tk.Toplevel(self.root)
        win.title("상품코드 기준 단가 변경")
        win.geometry("650x600")
        win.grab_set()

        top_frame = ttk.Frame(win, padding=10)
        top_frame.pack(fill='x')

        ttk.Label(top_frame, text="상품코드:").pack(side='left', padx=5)
        ent_icode = ttk.Entry(top_frame, width=20)
        ent_icode.pack(side='left', padx=5)

        cbo_mode = ttk.Combobox(top_frame, values=["정확히 일치", "부분 포함"], state="readonly", width=15)
        cbo_mode.current(0)
        cbo_mode.pack(side='left', padx=5)

        mid_frame = ttk.Frame(win, padding=10)
        mid_frame.pack(fill='both', expand=True)

        cols = ("바코드", "상품코드", "상품명", "현재 기본 단가", "현재 재고")
        tree_res = ttk.Treeview(mid_frame, columns=cols, show='headings')
        for col in cols:
            tree_res.heading(col, text=col)
            width = 120 if col == "상품명" else 100
            tree_res.column(col, width=width, anchor='center')
        tree_res.pack(fill='both', expand=True)

        def search_items(event=None):
            target_code = ent_icode.get().strip()
            search_mode = cbo_mode.get()

            if not target_code: return

            for row in tree_res.get_children():
                tree_res.delete(row)

            found_items = []
            for bcode, info in self.master_data.items():
                icode = info.get('item_code', '')
                if search_mode == "정확히 일치" and icode == target_code:
                    found_items.append((bcode, info))
                elif search_mode == "부분 포함" and target_code in icode:
                    found_items.append((bcode, info))

            if not found_items:
                messagebox.showinfo("안내", "해당 검색 조건에 해당하는 상품이 없습니다.", parent=win)
                lbl_count.config(text="현재 선택된 상품 수: 0 개")
                return

            for bcode, info in found_items:
                tree_res.insert("", tk.END, values=(bcode, info.get('item_code', ''), info['name'], info['price'], info['stock']))

            lbl_count.config(text=f"현재 선택된 상품 수: {len(found_items)} 개", foreground="blue")

        ttk.Button(top_frame, text="검색", command=search_items).pack(side='left', padx=10)
        ent_icode.bind('<Return>', search_items)

        bot_frame = ttk.Frame(win, padding=10)
        bot_frame.pack(fill='x')

        lbl_count = ttk.Label(bot_frame, text="현재 선택된 상품 수: 0 개", font=("Arial", 10, "bold"))
        lbl_count.grid(row=0, column=0, columnspan=2, pady=10, sticky='w')

        ttk.Label(bot_frame, text="새 기본 단가:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        ent_new_price = ttk.Entry(bot_frame, width=20)
        ent_new_price.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        ttk.Label(bot_frame, text="변경 메모:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        ent_memo = ttk.Entry(bot_frame, width=40)
        ent_memo.grid(row=2, column=1, padx=5, pady=5, sticky='w')

        def save_bulk_price_change():
            children = tree_res.get_children()
            if not children:
                messagebox.showwarning("경고", "단가를 변경할 상품을 먼저 검색해주세요.", parent=win)
                return

            try:
                new_price = int(ent_new_price.get().strip())
                if new_price < 0: raise ValueError
            except ValueError:
                messagebox.showerror("오류", "새 기본 단가는 0 이상의 숫자여야 합니다.", parent=win)
                return

            items_to_update = []
            for child in children:
                values = tree_res.item(child, 'values')
                bcode = str(values[0])
                old_price = int(values[3])
                if old_price != new_price:
                    items_to_update.append((bcode, old_price))

            if not items_to_update:
                messagebox.showinfo("안내", "변경할 단가가 없습니다. 모든 상품이 이미 같은 단가입니다.", parent=win)
                return

            confirm_msg = f"총 {len(items_to_update)}개 상품의 기본 단가를 {new_price:,}원으로 변경하시겠습니까?"
            if not messagebox.askyesno("확인", confirm_msg, parent=win):
                return

            try:
                wb = openpyxl.load_workbook(self.op_file_path)
                
                if "상품마스터" in wb.sheetnames:
                    ws_master = wb["상품마스터"]
                    update_dict = dict(items_to_update)
                    for row in ws_master.iter_rows(min_row=2):
                        bcode_in_excel = str(row[0].value) if row[0].value else ""
                        if bcode_in_excel in update_dict:
                            row[3].value = new_price

                if "단가변경이력" not in wb.sheetnames:
                    ws_hist = wb.create_sheet("단가변경이력")
                    ws_hist.append(["변경일자", "시간", "변경기준", "입력상품코드", "바코드", "상품코드", "상품명", "이전단가", "변경단가", "변경메모"])
                else:
                    ws_hist = wb["단가변경이력"]

                now = datetime.now()
                d_str = now.strftime("%Y-%m-%d")
                t_str = now.strftime("%H:%M:%S")
                mode_str = cbo_mode.get()
                input_code = ent_icode.get().strip()
                memo = ent_memo.get().strip()

                for bcode, old_price in items_to_update:
                    info = self.master_data[bcode]
                    ws_hist.append([d_str, t_str, mode_str, input_code, bcode, info.get('item_code',''), info['name'], old_price, new_price, memo])

                wb.save(self.op_file_path)

                for bcode, _ in items_to_update:
                    self.master_data[bcode]['price'] = new_price

                messagebox.showinfo("성공", f"총 {len(items_to_update)}개 상품의 기본 단가 변경이 완료되었습니다.", parent=win)
                search_items() 

            except PermissionError:
                messagebox.showerror("오류", "운영파일이 엑셀 등에서 열려 있습니다. 파일을 닫고 다시 시도해 주세요.", parent=win)
            except Exception as e:
                messagebox.showerror("오류", f"일괄 단가 변경 중 오류 발생:\n{e}", parent=win)

        ttk.Button(bot_frame, text="단가 변경 저장", command=save_bulk_price_change, style="Accent.TButton", width=30).grid(row=3, column=0, columnspan=2, pady=20)


    # ================= 5. 반품/교환 탭 =================
    def build_return_exchange_tab(self):
        top_frame = ttk.LabelFrame(self.tab_return_exchange, text="처리 유형", padding=10)
        top_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(top_frame, text="처리 구분:").grid(row=0, column=0, padx=5, pady=5)
        self.cbo_ret_type = ttk.Combobox(top_frame, values=["단순 반품", "반품교환"], state="readonly", width=15)
        self.cbo_ret_type.current(0)
        self.cbo_ret_type.grid(row=0, column=1, padx=5, pady=5)
        self.cbo_ret_type.bind('<<ComboboxSelected>>', self.toggle_exchange_fields)

        mid_frame = ttk.Frame(self.tab_return_exchange)
        mid_frame.pack(fill='both', expand=True, padx=10, pady=5)

        ret_frame = ttk.LabelFrame(mid_frame, text="[반품 대상 상품 정보]", padding=10)
        ret_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

        ttk.Label(ret_frame, text="바코드:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_ret_bcode = ttk.Entry(ret_frame, width=20)
        self.ent_ret_bcode.grid(row=0, column=1, padx=5, pady=5)
        self.ent_ret_bcode.bind('<Return>', lambda e: self.auto_fill_price(self.ent_ret_bcode, self.ent_ret_price, self.var_ret_name))
        self.ent_ret_bcode.bind('<FocusOut>', lambda e: self.auto_fill_price(self.ent_ret_bcode, self.ent_ret_price, self.var_ret_name))

        self.var_ret_name = tk.StringVar(value="(상품명 대기)")
        ttk.Label(ret_frame, textvariable=self.var_ret_name, foreground="blue").grid(row=0, column=2, padx=5, pady=5, sticky='w')

        ttk.Label(ret_frame, text="시리얼:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ent_ret_serial = ttk.Entry(ret_frame, width=20)
        self.ent_ret_serial.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(ret_frame, text="수량:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.ent_ret_qty = ttk.Entry(ret_frame, width=20)
        self.ent_ret_qty.insert(0, "1")
        self.ent_ret_qty.grid(row=2, column=1, padx=5, pady=5)
        self.ent_ret_qty.bind('<KeyRelease>', self.calculate_diff)

        ttk.Label(ret_frame, text="단가:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.ent_ret_price = ttk.Entry(ret_frame, width=20)
        self.ent_ret_price.grid(row=3, column=1, padx=5, pady=5)
        self.ent_ret_price.bind('<KeyRelease>', self.calculate_diff)

        self.exc_frame = ttk.LabelFrame(mid_frame, text="[교환 출고 상품 정보]", padding=10)
        self.exc_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        ttk.Label(self.exc_frame, text="바코드:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_exc_bcode = ttk.Entry(self.exc_frame, width=20, state='disabled')
        self.ent_exc_bcode.grid(row=0, column=1, padx=5, pady=5)
        self.ent_exc_bcode.bind('<Return>', lambda e: self.auto_fill_price(self.ent_exc_bcode, self.ent_exc_price, self.var_exc_name))
        self.ent_exc_bcode.bind('<FocusOut>', lambda e: self.auto_fill_price(self.ent_exc_bcode, self.ent_exc_price, self.var_exc_name))

        self.var_exc_name = tk.StringVar(value="(상품명 대기)")
        ttk.Label(self.exc_frame, textvariable=self.var_exc_name, foreground="blue").grid(row=0, column=2, padx=5, pady=5, sticky='w')

        ttk.Label(self.exc_frame, text="시리얼:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ent_exc_serial = ttk.Entry(self.exc_frame, width=20, state='disabled')
        self.ent_exc_serial.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(self.exc_frame, text="수량:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.ent_exc_qty = ttk.Entry(self.exc_frame, width=20, state='disabled')
        self.ent_exc_qty.insert(0, "1")
        self.ent_exc_qty.grid(row=2, column=1, padx=5, pady=5)
        self.ent_exc_qty.bind('<KeyRelease>', self.calculate_diff)

        ttk.Label(self.exc_frame, text="단가:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.ent_exc_price = ttk.Entry(self.exc_frame, width=20, state='disabled')
        self.ent_exc_price.grid(row=3, column=1, padx=5, pady=5)
        self.ent_exc_price.bind('<KeyRelease>', self.calculate_diff)

        bot_frame = ttk.Frame(self.tab_return_exchange)
        bot_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(bot_frame, text="처리 메모:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_memo = ttk.Entry(bot_frame, width=50)
        self.ent_memo.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        self.lbl_diff = ttk.Label(bot_frame, text="차액: 0 원", font=("Arial", 12, "bold"))
        self.lbl_diff.grid(row=1, column=0, columnspan=2, pady=10, sticky='w', padx=5)

        self.btn_process_ret = ttk.Button(bot_frame, text="반품 / 교환 처리 실행", command=self.process_return_exchange, style="Accent.TButton", width=30)
        self.btn_process_ret.grid(row=2, column=0, columnspan=2, pady=10)

    def toggle_exchange_fields(self, event=None):
        mode = self.cbo_ret_type.get()
        state = 'normal' if mode == "반품교환" else 'disabled'
        
        self.ent_exc_bcode.config(state=state)
        self.ent_exc_serial.config(state=state)
        self.ent_exc_qty.config(state=state)
        self.ent_exc_price.config(state=state)
        
        self.calculate_diff()

    def auto_fill_price(self, bcode_entry, price_entry, name_var):
        bcode = bcode_entry.get().strip()
        if bcode and bcode in self.master_data:
            name_var.set(self.master_data[bcode]['name'])
            price_entry.config(state='normal')
            price_entry.delete(0, tk.END)
            price_entry.insert(0, str(self.master_data[bcode]['price']))
            if self.cbo_ret_type.get() == "단순 반품" and bcode_entry == self.ent_exc_bcode:
                price_entry.config(state='disabled')
            self.calculate_diff()
        elif bcode:
            name_var.set("(미등록 바코드)")

    def calculate_diff(self, event=None):
        try:
            ret_qty = int(self.ent_ret_qty.get() or 0)
            ret_price = int(self.ent_ret_price.get() or 0)
            ret_amt = ret_qty * ret_price
            
            exc_amt = 0
            if self.cbo_ret_type.get() == "반품교환":
                exc_qty = int(self.ent_exc_qty.get() or 0)
                exc_price = int(self.ent_exc_price.get() or 0)
                exc_amt = exc_qty * exc_price

            diff = exc_amt - ret_amt
            if diff > 0:
                text = f"차액: {diff:,} 원 (고객 추가 결제 필요)"
                color = "red"
            elif diff < 0:
                text = f"차액: {diff:,} 원 (고객 환불 필요)"
                color = "blue"
            else:
                text = "차액: 0 원"
                color = "black"

            self.lbl_diff.config(text=text, foreground=color)
        except ValueError:
            self.lbl_diff.config(text="차액: 입력값 오류", foreground="gray")

    def validate_returnable_item(self, wb, barcode, qty, serial):
        if "판매상세" not in wb.sheetnames:
            return False, "판매 이력(판매상세 시트)이 존재하지 않습니다."
            
        ws_sales = wb["판매상세"]
        headers = {cell.value: idx for idx, cell in enumerate(ws_sales[1]) if cell.value}
        
        if '바코드' not in headers or '수량' not in headers:
            return False, "판매상세 시트의 헤더 정보를 찾을 수 없습니다."
            
        barcode_idx = headers['바코드']
        qty_idx = headers['수량']
        serial_idx = headers.get('시리얼코드', -1)
        
        net_qty_total = 0
        net_qty_serial = 0
        serial_found = False
        
        for row in ws_sales.iter_rows(min_row=2, values_only=True):
            row_barcode = str(row[barcode_idx]) if row[barcode_idx] is not None else ""
            if row_barcode == barcode:
                try:
                    row_qty = int(row[qty_idx] or 0)
                except ValueError:
                    row_qty = 0
                
                net_qty_total += row_qty
                
                if serial and serial_idx != -1:
                    row_serial = str(row[serial_idx]) if row[serial_idx] is not None else ""
                    if row_serial == serial:
                        serial_found = True
                        net_qty_serial += row_qty
                        
        if serial:
            if not serial_found:
                return False, "해당 시리얼의 판매 이력을 찾을 수 없습니다."
            if net_qty_serial <= 0:
                return False, "이미 반품 처리된 시리얼입니다."
        else:
            if net_qty_total <= 0:
                return False, "판매 이력이 없는 상품입니다. 반품/교환 처리할 수 없습니다."
            if net_qty_total < qty:
                return False, f"반품 가능 수량을 초과했습니다. 현재 반품 가능 수량: {net_qty_total}개"
                
        return True, ""

    def process_return_exchange(self):
        if not self.op_file_path or not self.master_data:
            messagebox.showwarning("경고", "기준 엑셀 및 운영파일 설정이 필요합니다.")
            return

        ret_type = self.cbo_ret_type.get()
        memo = self.ent_memo.get().strip()

        ret_bcode = self.ent_ret_bcode.get().strip()
        ret_serial = self.ent_ret_serial.get().strip()
        try:
            ret_qty = int(self.ent_ret_qty.get() or 0)
            ret_price = int(self.ent_ret_price.get() or 0)
        except ValueError:
            messagebox.showerror("오류", "수량과 단가는 숫자로 입력해야 합니다.")
            return

        if not ret_bcode or ret_bcode not in self.master_data:
            messagebox.showerror("오류", "반품 대상 바코드가 유효하지 않습니다.")
            return
        if ret_qty <= 0:
            messagebox.showerror("오류", "반품 수량은 1 이상이어야 합니다.")
            return

        if self.master_data[ret_bcode]['category'] in ['RACKET', 'RUBBER']:
            if not ret_serial:
                messagebox.showerror("오류", "라켓/러버 상품은 반품 시 시리얼 입력이 필수입니다.")
                return

        ret_amt = ret_qty * ret_price
        ret_name = self.master_data[ret_bcode]['name']

        exc_bcode = ""
        exc_serial = ""
        exc_qty = 0
        exc_price = 0
        exc_amt = 0
        exc_name = ""

        if ret_type == "반품교환":
            exc_bcode = self.ent_exc_bcode.get().strip()
            exc_serial = self.ent_exc_serial.get().strip()
            try:
                exc_qty = int(self.ent_exc_qty.get() or 0)
                exc_price = int(self.ent_exc_price.get() or 0)
            except ValueError:
                messagebox.showerror("오류", "교환 수량과 단가는 숫자로 입력해야 합니다.")
                return

            if not exc_bcode or exc_bcode not in self.master_data:
                messagebox.showerror("오류", "교환 출고 대상 바코드가 유효하지 않습니다.")
                return
            if exc_qty <= 0:
                messagebox.showerror("오류", "교환 수량은 1 이상이어야 합니다.")
                return

            if self.master_data[exc_bcode]['stock'] < exc_qty:
                messagebox.showerror("재고 부족", f"[{self.master_data[exc_bcode]['name']}]의 전산 재고가 부족합니다. (현재: {self.master_data[exc_bcode]['stock']}개)")
                return

            is_exc_serial_req = self.master_data[exc_bcode]['category'] in ['RUBBER', 'RACKET']
            if is_exc_serial_req and not exc_serial:
                messagebox.showerror("오류", "교환 출고 상품은 시리얼 입력이 필수입니다.")
                return
            if hasattr(self, 'sold_serial_data') and exc_bcode in self.sold_serial_data and exc_serial in self.sold_serial_data[exc_bcode]:
                messagebox.showerror("오류", "교환 출고하려는 시리얼 번호가 이미 다른 거래에서 판매 완료되었습니다.")
                return

            exc_amt = exc_qty * exc_price
            exc_name = self.master_data[exc_bcode]['name']

        diff = exc_amt - ret_amt
        confirm_msg = f"처리유형: {ret_type}\n\n[반품] {ret_name} ({ret_qty}개)\n"
        if ret_type == "반품교환":
            confirm_msg += f"[교환] {exc_name} ({exc_qty}개)\n\n차액: {diff:,}원\n"
        confirm_msg += "\n위 내용으로 처리를 진행하시겠습니까?"

        if not messagebox.askyesno("확인", confirm_msg):
            return

        try:
            wb = openpyxl.load_workbook(self.op_file_path)
            
            is_valid, err_msg = self.validate_returnable_item(wb, ret_bcode, ret_qty, ret_serial)
            if not is_valid:
                wb.close()
                messagebox.showerror("반품 불가", err_msg)
                return

            ws_sales = wb["판매상세"]
            ws_ret_exc = wb["반품교환상세"]
            ws_stock = wb["재고현황"]
            ws_serial = wb["시리얼목록"]
            ws_daily = wb["일별집계"]

            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")

            # 1. 반품 처리 (메모리 반영)
            self.master_data[ret_bcode]['stock'] += ret_qty
            if ret_serial:
                if hasattr(self, 'sold_serial_data') and ret_bcode in self.sold_serial_data and ret_serial in self.sold_serial_data[ret_bcode]:
                    self.sold_serial_data[ret_bcode].remove(ret_serial)
                if ret_bcode not in self.serial_data:
                    self.serial_data[ret_bcode] = []
                if ret_serial not in self.serial_data[ret_bcode]:
                    self.serial_data[ret_bcode].append(ret_serial)

            # 1. 반품 기록 (엑셀)
            ws_sales.append([date_str, time_str, ret_bcode, ret_name, -ret_qty, ret_price, -ret_amt, ret_serial, "반품" if ret_type == "단순 반품" else "교환반품"])
            self._update_daily_report_excel(ws_daily, date_str, ret_bcode, ret_name, -ret_qty, -ret_amt)
            self._update_serial_status_excel(ws_serial, ret_bcode, ret_serial, status="반품")

            # 2. 교환 처리 (메모리 반영)
            if ret_type == "반품교환":
                self.master_data[exc_bcode]['stock'] -= exc_qty
                if exc_serial:
                    if exc_bcode in self.serial_data and exc_serial in self.serial_data[exc_bcode]:
                        self.serial_data[exc_bcode].remove(exc_serial)
                    if not hasattr(self, 'sold_serial_data'): self.sold_serial_data = {}
                    if exc_bcode not in self.sold_serial_data: self.sold_serial_data[exc_bcode] = set()
                    self.sold_serial_data[exc_bcode].add(exc_serial)

                # 2. 교환 기록 (엑셀)
                ws_sales.append([date_str, time_str, exc_bcode, exc_name, exc_qty, exc_price, exc_amt, exc_serial, "교환출고"])
                self._update_daily_report_excel(ws_daily, date_str, exc_bcode, exc_name, exc_qty, exc_amt)
                self._update_serial_status_excel(ws_serial, exc_bcode, exc_serial, status="O", add_if_not_exist=True, item_name=exc_name)

            # 3. 반품교환상세 시트 기록
            ws_ret_exc.append([date_str, time_str, ret_type, "", ret_bcode, ret_name, ret_qty, ret_price, ret_amt, ret_serial, 
                               exc_bcode, exc_name, exc_qty, exc_price, exc_amt, exc_serial, diff, memo])

            # 4. 재고현황 최신화
            ws_stock.delete_rows(2, ws_stock.max_row)
            for b_code, b_data in self.master_data.items():
                ws_stock.append([b_code, b_data['name'], b_data['stock']])

            wb.save(self.op_file_path)

            msg = "처리가 완료되었습니다."
            if ret_type == "반품교환":
                if diff > 0: msg += f"\n\n※ 고객님께 {diff:,}원을 추가 결제 받아주세요."
                elif diff < 0: msg += f"\n\n※ 고객님께 {-diff:,}원을 환불해 주세요."
            messagebox.showinfo("완료", msg)
            
            # 입력창 초기화
            self.ent_ret_bcode.delete(0, tk.END)
            self.ent_ret_serial.delete(0, tk.END)
            self.ent_ret_price.delete(0, tk.END)
            self.ent_exc_bcode.delete(0, tk.END)
            self.ent_exc_serial.delete(0, tk.END)
            self.ent_exc_price.delete(0, tk.END)
            self.ent_memo.delete(0, tk.END)
            self.var_ret_name.set("(상품명 대기)")
            self.var_exc_name.set("(상품명 대기)")
            self.calculate_diff()

        except PermissionError:
            self._rollback_memory(ret_type, ret_bcode, ret_qty, ret_serial, exc_bcode, exc_qty, exc_serial)
            messagebox.showerror("오류", "운영파일이 엑셀 등에서 열려 있습니다. 파일을 닫고 다시 시도해 주세요.")
        except Exception as e:
            self._rollback_memory(ret_type, ret_bcode, ret_qty, ret_serial, exc_bcode, exc_qty, exc_serial)
            messagebox.showerror("오류", f"처리 중 오류 발생 (데이터 복구됨):\n{e}")

    def _update_daily_report_excel(self, ws_daily, date_str, bcode, name, qty, amt):
        found = False
        for row in ws_daily.iter_rows(min_row=2):
            if str(row[0].value)[:10] == date_str and str(row[1].value) == bcode:
                row[3].value = int(row[3].value or 0) + qty
                row[4].value = int(row[4].value or 0) + amt
                found = True
                break
        if not found:
            ws_daily.append([date_str, bcode, name, qty, amt])

    def _update_serial_status_excel(self, ws_serial, bcode, serial, status, add_if_not_exist=False, item_name=""):
        if not serial: return
        found = False
        for row in ws_serial.iter_rows(min_row=2):
            if str(row[0].value) == bcode and str(row[2].value) == serial:
                row[3].value = status if status != "반품" else "" 
                found = True
                break
        if not found and add_if_not_exist:
            ws_serial.append([bcode, item_name, serial, status])

    def _rollback_memory(self, ret_type, ret_bcode, ret_qty, ret_serial, exc_bcode, exc_qty, exc_serial):
        if ret_bcode in self.master_data:
            self.master_data[ret_bcode]['stock'] -= ret_qty
        if ret_serial:
            if ret_bcode in self.serial_data and ret_serial in self.serial_data[ret_bcode]:
                self.serial_data[ret_bcode].remove(ret_serial)
            if hasattr(self, 'sold_serial_data') and ret_bcode in self.sold_serial_data:
                self.sold_serial_data[ret_bcode].add(ret_serial)
        
        if ret_type == "반품교환":
            if exc_bcode in self.master_data:
                self.master_data[exc_bcode]['stock'] += exc_qty
            if exc_serial:
                if hasattr(self, 'sold_serial_data') and exc_bcode in self.sold_serial_data and exc_serial in self.sold_serial_data[exc_bcode]:
                    self.sold_serial_data[exc_bcode].remove(exc_serial)
                if exc_bcode not in self.serial_data: self.serial_data[exc_bcode] = []
                self.serial_data[exc_bcode].append(exc_serial)

if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style()
    style.theme_use('clam')
    app = BoothSalesApp(root)
    root.mainloop()
